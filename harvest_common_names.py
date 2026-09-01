#!/usr/bin/env python3
"""
harvest_common_names.py — auto-fetch English common names for POPPy plants that lack one.

Replicates the manual "google each plant" pass, automatically, from two public APIs that
return vernacular names WITH a source you can cite:
  * GBIF   — /species/match  ->  /species/{key}/vernacularNames   (primary)
  * iNaturalist — /v1/taxa?q=<name>&locale=en  (fallback: preferred_common_name)

Output: an .xlsx review sheet in the SAME columns as the file you curated, containing only the
plants where a name was found, so you can review/edit and keep the good ones. Nothing is written
back to the ontology — this is a proposal for your review.

Run it in your networked Terminal (the conda env that runs your GBIF pipeline):

    python3 harvest_common_names.py \
        --plants-index website/data/plants_index.json \
        --out data/enrichment/common_names_auto_review.xlsx

Resumable: progress is cached in <out>.cache.json — re-run to continue after a stop.
Deps: requests, openpyxl (already in your pipeline env).
"""
from __future__ import annotations
import argparse, json, os, sys, time, re
from urllib.parse import quote

import requests
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

GBIF_MATCH = "https://api.gbif.org/v1/species/match?name={name}"
GBIF_VERN  = "https://api.gbif.org/v1/species/{key}/vernacularNames"
GBIF_SPECIES_URL = "https://www.gbif.org/species/{key}"
INAT = "https://api.inaturalist.org/v1/taxa?q={name}&locale=en&per_page=1"

ENGLISH = {"eng", "en"}
HEADERS = {"User-Agent": "poppy-commonname-harvest/1.0 (research; contact bullseye075@gmail.com)"}

def get(url, tries=3, timeout=20):
    for i in range(tries):
        try:
            r = requests.get(url, headers=HEADERS, timeout=timeout)
            if r.status_code == 200:
                return r.json()
            if r.status_code in (429, 502, 503):
                time.sleep(2 * (i + 1)); continue
            return None
        except requests.RequestException:
            time.sleep(1.5 * (i + 1))
    return None

def clean(nm: str) -> str:
    # strip author citations / rank tags to improve match rate: keep first 2-3 tokens
    nm = re.sub(r"\s*\(.*?\)", "", nm)                    # drop parentheticals
    nm = re.sub(r"\b(var|subsp|ssp|f|cv)\.?\s+\w+.*$", "", nm)  # drop infraspecific tails
    return nm.strip()

def gbif_names(sciname: str, max_other=5):
    """Return (english_names, other_names_tagged, source_url).

    english_names: plain English vernaculars.
    other_names_tagged: up to `max_other` non-English names formatted "name [lang]"
    (you asked to keep non-English/transliterated names too, like your Malay/pinyin entries).
    """
    m = get(GBIF_MATCH.format(name=quote(sciname)))
    if not m or "usageKey" not in m:
        m = get(GBIF_MATCH.format(name=quote(clean(sciname))))
    if not m or "usageKey" not in m:
        return [], [], None
    key = m.get("acceptedUsageKey") or m["usageKey"]
    v = get(GBIF_VERN.format(key=key))
    if not v:
        return [], [], GBIF_SPECIES_URL.format(key=key)
    eng, other = [], []
    for rec in v.get("results", []):
        nm = (rec.get("vernacularName") or "").strip()
        if not nm:
            continue
        lang = (rec.get("language") or "").lower()
        if lang in ENGLISH:
            eng.append(nm)
        else:
            other.append(f"{nm} [{lang or '?'}]")
    def dedup(xs):
        seen, out = set(), []
        for x in xs:
            k = x.lower()
            if k not in seen:
                seen.add(k); out.append(x)
        return out
    return dedup(eng), dedup(other)[:max_other], GBIF_SPECIES_URL.format(key=key)

def inat_name(sciname: str):
    j = get(INAT.format(name=quote(sciname)))
    if not j:
        return None, None
    for res in j.get("results", []):
        cn = res.get("preferred_common_name")
        if cn:
            return cn, f"https://www.inaturalist.org/taxa/{res.get('id')}"
    return None, None

def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--plants-index", required=True)
    ap.add_argument("--out", default="common_names_auto_review.xlsx")
    ap.add_argument("--limit", type=int, default=0, help="max plants to process (0=all)")
    ap.add_argument("--sleep", type=float, default=0.15, help="seconds between plants")
    args = ap.parse_args()

    data = json.load(open(args.plants_index))
    todo = [r for r in data if not r.get("common")]
    todo.sort(key=lambda r: -(r.get("nc", 0) or 0))     # richest plants first
    if args.limit:
        todo = todo[: args.limit]

    cache_path = args.out + ".cache.json"
    cache = {}
    if os.path.exists(cache_path):
        cache = json.load(open(cache_path))
    print(f"{len(todo)} unnamed plants; {len(cache)} already in cache", file=sys.stderr)

    for i, r in enumerate(todo, 1):
        nm = r["name"]
        if nm in cache:
            continue
        eng, other, src = gbif_names(nm)
        source = src
        # prefer English; if none, try iNaturalist's English common name; else keep GBIF's other-language names
        if eng:
            names = eng
        else:
            cn, isrc = inat_name(nm)
            if cn:
                names, source = [cn], isrc
            else:
                names = other        # non-English fallback, tagged with [lang]
        cache[nm] = {"common": "; ".join(names), "source": source or "",
                     "english": bool(eng)}
        if i % 25 == 0:
            json.dump(cache, open(cache_path, "w"), ensure_ascii=False)
            print(f"  {i}/{len(todo)}  (found so far: {sum(1 for v in cache.values() if v['common'])})",
                  file=sys.stderr)
        time.sleep(args.sleep)
    json.dump(cache, open(cache_path, "w"), ensure_ascii=False)

    # build review workbook: only plants where a name was found
    wb = Workbook(); ws = wb.active; ws.title = "Auto common names (review)"
    ws.append(["Scientific name", "Compounds (nc)", "Trials", "Organism ID",
               "Common Name (proposed)", "Source"])
    hf = Font(name="Arial", bold=True, color="FFFFFF"); hb = PatternFill("solid", fgColor="2F5D3A")
    for c in ws[1]: c.font = hf; c.fill = hb
    found = [r for r in todo if cache.get(r["name"], {}).get("common")]
    found.sort(key=lambda r: -(r.get("nc", 0) or 0))
    body = Font(name="Arial")
    for r in found:
        c = cache[r["name"]]
        ws.append([r["name"], r.get("nc", 0) or 0, r.get("trials", 0) or 0, r["id"],
                   c["common"], c["source"]])
    for row in ws.iter_rows(min_row=2):
        for cell in row: cell.font = body
    for i, w in enumerate([46, 14, 10, 16, 46, 60], 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A2"; ws.auto_filter.ref = f"A1:F{ws.max_row}"
    wb.save(args.out)
    print(f"\nDONE. {len(found)} of {len(todo)} unnamed plants got a proposed common name.")
    print(f"Review sheet: {args.out}")

if __name__ == "__main__":
    raise SystemExit(main())
